from openpyxl import load_workbook


wb = load_workbook("./test.xlsx")


# 1.获取sheet
print(wb.sheetnames)

# # 2.选择sheet
# sheet = wb["Sheet1"]
# 3 选择sheet，基于索引位置
sheet = wb.worksheets[0]

cell = sheet.cell(1, 1)  # 第一行第一列

# print(dir(sheet))

print(cell.value,cell.row, dir(cell))


# 4. 循环所以的sheet
for name in wb.sheetnames:
    print(name)
    sheet = wb[name]
    cell = sheet.cell(1, 1)
    print(cell.value)

for sheet in wb.worksheets:
    print(sheet)
    cell = sheet.cell(1,1)
    print(cell.value)

for sheet in wb:
    print(sheet)
    cell = sheet.cell(1,1)
    print(cell.value)