from openpyxl import load_workbook
from openpyxl import workbook


# 1.在原来的excel上写
'''
wb = load_workbook("./site_list.xlsx")
sheet = wb.worksheets[0]

# 找到单元格

cell = sheet.cell(4,1)
cell.value = "new write"

wb.save("./site_list.xlsx")
'''
# 2.在新创建的excel上写

# 创建一个sheet
wb = workbook.Workbook()

print(wb.sheetnames)

# wb.create_sheet("test_sheet1")

sheet = wb.worksheets[0]   # 或这个 sheet = wb["Sheet"]

sheet.title = "test_sheet"
print(wb.sheetnames)

cell = sheet.cell(1, 1)
cell.value = '新的开始'


# 删除
# del wb['test_sheet1']

wb.save('./create.xlsx')


## 获取某些单元格
'''
cell_list = sheet["B2": "C3"]

'''
