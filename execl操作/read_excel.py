from openpyxl import load_workbook
wb = load_workbook("./site_list.xlsx")

# wb = load_workbook("./test_xls.xls")


sheet = wb.worksheets[0]
# 1. 获取第n行n列
'''
cell = sheet.cell(1,1)

print(cell.value)
print(cell.style)
print(cell.font)
print(cell.alignment)
'''

# 2. 获取某个单元格
'''
c1 = sheet["C1"]
print(c1.value)

c2 = sheet["C2"]
print(c2.value)
'''

# 3.获取第n行所有的单元格
row = sheet[1]
print(row)
for cell in row:
    print(cell.value)


# 4.获取所有行的数据
for row in sheet.rows:
    print(row[0].value) # 第一列的数据


# 5. 获取所有的列数据
for col in sheet.columns:
    print(col[0].value) # 第一行的数据


## 读取合并的单元格
## 第一个的值Cell，其他的为空MergedCell


